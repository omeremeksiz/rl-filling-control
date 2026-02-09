# scripts/eval_compare_test.py
from __future__ import annotations

import json
import logging
import os
import re
from datetime import datetime
from typing import Any, Dict, Iterable, List, Mapping, Optional, Tuple

import numpy as np
import openpyxl
import yaml
import math

from utils.plotting_utils import (
    plot_summary_switching_trajectory,
    plot_multi_qvalue_vs_state,
    plot_multi_qvalue_pair_tables,
)

CONFIG_PATH = os.path.join("configs", "eval_compare_test.yaml")

_EPISODE_RE = re.compile(r"--- Episode (\d+)/(\d+) ---")
_MODEL_RE = re.compile(r"Model-Selected Next Switching Point:\s*(.+)")
_EXPLORE_RE = re.compile(r"Explored Switching Point:\s*(.+)")
_EXPLORE_META_RE = re.compile(r"Explored:\s*([^\n]+)", re.IGNORECASE)


def load_config(path: str = CONFIG_PATH) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as handle:
        return yaml.safe_load(handle)


def _resolve_plot_format(value: Any, default: str = "png") -> str:
    if value is None:
        return default
    fmt = str(value).strip().lower().lstrip(".")
    if fmt not in {"pdf", "png"}:
        return default
    return fmt


def _resolve_bool(value: Any, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "yes", "1", "on"}:
            return True
        if lowered in {"false", "no", "0", "off"}:
            return False
    return default


def _parse_log_trajectory(log_path: str, max_episodes: Optional[int]) -> Tuple[List[int], List[int], List[Optional[int]]]:
    model_selected: Dict[int, int] = {}
    explored: Dict[int, Optional[int]] = {}
    current_episode: Optional[int] = None
    max_seen = 0

    with open(log_path, "r", encoding="utf-8") as handle:
        for line in handle:
            match = _EPISODE_RE.search(line)
            if match:
                current_episode = int(match.group(1))
                max_seen = max(max_seen, current_episode)
                continue
            if current_episode is None:
                continue
            model_match = _MODEL_RE.search(line)
            if model_match:
                value = model_match.group(1).strip()
                try:
                    model_selected[current_episode] = int(float(value))
                except ValueError:
                    pass
                continue
            explore_match = _EXPLORE_RE.search(line)
            if explore_match:
                value = explore_match.group(1).strip()
                if value.lower() == "none":
                    explored[current_episode] = None
                else:
                    try:
                        explored[current_episode] = int(float(value))
                    except ValueError:
                        explored[current_episode] = None

    episode_limit = min(max_seen, max_episodes) if max_episodes else max_seen
    episodes: List[int] = []
    model_list: List[int] = []
    explored_list: List[Optional[int]] = []
    for ep in range(1, episode_limit + 1):
        if ep not in model_selected:
            continue
        episodes.append(ep)
        model_list.append(model_selected[ep])
        explored_list.append(explored.get(ep))
    return episodes, model_list, explored_list


def _count_excel_episodes(ws, columns_per_episode: int) -> int:
    count = 0
    col = 1
    while True:
        value = ws.cell(row=1, column=col).value
        if value is None:
            break
        if str(value).strip().lower().startswith("episode"):
            count += 1
        col += columns_per_episode
    return count


def _load_explored_series_from_excel(
    excel_path: str,
    columns_per_episode: int,
    *,
    max_episodes: Optional[int],
) -> List[Optional[int]]:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    total_eps = _count_excel_episodes(ws, columns_per_episode)
    limit = min(total_eps, max_episodes) if max_episodes else total_eps
    explored_series: List[Optional[int]] = []
    for idx in range(1, limit + 1):
        start_col = 1 + (idx - 1) * columns_per_episode
        meta_cell = ws.cell(row=2, column=start_col + 2).value
        if not meta_cell:
            explored_series.append(None)
            continue
        match = _EXPLORE_META_RE.search(str(meta_cell))
        if not match:
            explored_series.append(None)
            continue
        value = match.group(1).strip()
        if value.lower() == "none":
            explored_series.append(None)
        else:
            try:
                explored_series.append(int(float(value)))
            except ValueError:
                explored_series.append(None)
    wb.close()
    return explored_series


def _load_mab_q_table(excel_path: str, episode_num: int) -> Dict[int, float]:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    max_episodes = _count_excel_episodes(ws, 3)
    episode_idx = min(max(1, episode_num), max_episodes) if max_episodes else episode_num
    start_col = 1 + (episode_idx - 1) * 3

    q_table: Dict[int, float] = {}
    row = 4
    while True:
        sp = ws.cell(row=row, column=start_col).value
        if sp is None:
            break
        q_val = ws.cell(row=row, column=start_col + 2).value
        try:
            q_table[int(sp)] = float(q_val)
        except (TypeError, ValueError):
            pass
        row += 1
    wb.close()
    return q_table


def _load_mc_q_table(excel_path: str, episode_num: int) -> Dict[Tuple[int, int], float]:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    max_episodes = _count_excel_episodes(ws, 4)
    episode_idx = min(max(1, episode_num), max_episodes) if max_episodes else episode_num
    start_col = 1 + (episode_idx - 1) * 4

    q_table: Dict[Tuple[int, int], float] = {}
    row = 4
    while True:
        weight = ws.cell(row=row, column=start_col).value
        if weight is None:
            break
        action = ws.cell(row=row, column=start_col + 1).value
        q_val = ws.cell(row=row, column=start_col + 3).value
        try:
            q_table[(int(weight), int(action))] = float(q_val)
        except (TypeError, ValueError):
            pass
        row += 1
    wb.close()
    return q_table


def _infer_tick_step(states: Iterable[int], max_labels: int = 15) -> Optional[int]:
    state_list = sorted({int(val) for val in states})
    if not state_list:
        return None
    if len(state_list) <= max_labels:
        return 1
    span = state_list[-1] - state_list[0]
    if span <= 0:
        return 1
    return max(1, int(math.ceil(span / max(1, max_labels - 1))))


def main() -> None:
    cfg = load_config()
    outputs_dir = cfg.get("outputs_dir", "outputs")
    plot_cfg = cfg.get("plot", {})
    episodes_to_plot = plot_cfg.get("episodes")
    qvalue_episode = plot_cfg.get("qvalue_episode")
    compare_tick_fontsize = int(plot_cfg.get("compare_tick_fontsize", 32))
    qvalue_tick_fontsize = int(plot_cfg.get("qvalue_tick_fontsize", 28))
    show_exploration = _resolve_bool(cfg.get("plot_show_exploration"), default=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.join(outputs_dir, f"eval_compare_test_{timestamp}")
    os.makedirs(output_dir, exist_ok=True)
    qvalues_dir = os.path.join(output_dir, "qvalues")
    os.makedirs(qvalues_dir, exist_ok=True)

    log_path = os.path.join(output_dir, "eval_compare_test.log")
    logger = logging.getLogger(f"eval_compare_test_{timestamp}")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    handler = logging.FileHandler(log_path)
    handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    logger.addHandler(handler)
    logger.propagate = False

    plot_format = _resolve_plot_format(cfg.get("plot_format"), default="png")

    outputs_cfg = cfg.get("outputs", {})
    mab_runs = outputs_cfg.get("mab_runs", [])
    mc_runs = outputs_cfg.get("mc_runs", [])

    trajectory_data: Dict[str, Tuple[List[int], List[float]]] = {}
    exploration_data: Dict[str, List[Optional[float]]] = {}
    exploration_colors: Dict[str, str] = {}
    best_sp_mab: Dict[str, Optional[int]] = {}
    best_sp_mc: Dict[str, Optional[int]] = {}
    qvalue_tables_mab: Dict[str, Dict[int, float]] = {}
    qvalue_tables_mc: Dict[str, Dict[Tuple[int, int], float]] = {}

    def resolve_q_episode(total: int) -> int:
        if qvalue_episode is None:
            return total
        try:
            return min(max(1, int(qvalue_episode)), total)
        except (TypeError, ValueError):
            return total

    for run in mab_runs:
        run_path = run.get("path")
        label = run.get("label") or os.path.basename(run_path)
        if not run_path:
            continue
        log_file = os.path.join(run_path, "training_process.log")
        episodes, model_list, explored_list = _parse_log_trajectory(log_file, episodes_to_plot)
        if not episodes:
            logger.warning("No episodes parsed for MAB run %s", run_path)
            continue
        plot_label = f"MAB - {label}"
        trajectory_data[plot_label] = (episodes, [float(val) for val in model_list])
        best_sp_mab[plot_label] = model_list[-1] if model_list else None
        if explored_list and any(val is not None for val in explored_list):
            exploration_data[plot_label] = explored_list
        else:
            excel_path = os.path.join(run_path, "mab_qvalue_updates.xlsx")
            if os.path.exists(excel_path):
                exploration_data[plot_label] = _load_explored_series_from_excel(
                    excel_path,
                    3,
                    max_episodes=len(episodes),
                )
            else:
                exploration_data[plot_label] = explored_list
        exploration_colors[plot_label] = "#FF9100"

        excel_path = os.path.join(run_path, "mab_qvalue_updates.xlsx")
        if os.path.exists(excel_path):
            q_ep = resolve_q_episode(len(episodes))
            qvalue_tables_mab[plot_label] = _load_mab_q_table(excel_path, q_ep)

    for run in mc_runs:
        run_path = run.get("path")
        label = run.get("label") or os.path.basename(run_path)
        if not run_path:
            continue
        log_file = os.path.join(run_path, "training_process.log")
        episodes, model_list, explored_list = _parse_log_trajectory(log_file, episodes_to_plot)
        if not episodes:
            logger.warning("No episodes parsed for MC run %s", run_path)
            continue
        plot_label = f"MC - {label}"
        trajectory_data[plot_label] = (episodes, [float(val) for val in model_list])
        best_sp_mc[plot_label] = model_list[-1] if model_list else None
        if explored_list and any(val is not None for val in explored_list):
            exploration_data[plot_label] = explored_list
        else:
            excel_path = os.path.join(run_path, "mc_qvalue_updates.xlsx")
            if os.path.exists(excel_path):
                exploration_data[plot_label] = _load_explored_series_from_excel(
                    excel_path,
                    4,
                    max_episodes=len(episodes),
                )
            else:
                exploration_data[plot_label] = explored_list
        exploration_colors[plot_label] = "#00E5FF"

        excel_path = os.path.join(run_path, "mc_qvalue_updates.xlsx")
        if os.path.exists(excel_path):
            q_ep = resolve_q_episode(len(episodes))
            qvalue_tables_mc[plot_label] = _load_mc_q_table(excel_path, q_ep)

    compare_path = os.path.join(output_dir, f"compare_switching_trajectory.{plot_format}")
    plot_summary_switching_trajectory(
        trajectory_data,
        compare_path,
        explored_series=exploration_data,
        exploration_colors=exploration_colors,
        show_legend=False,
        show_titles=False,
        tick_fontsize=compare_tick_fontsize,
        show_exploration=show_exploration,
    )

    for label, table in qvalue_tables_mab.items():
        q_path = os.path.join(qvalues_dir, f"{label.replace(' ', '_').lower()}_qvalues.{plot_format}")
        states = list(table.keys())
        state_min = min(states) if states else None
        state_max = max(states) if states else None
        state_step = _infer_tick_step(states)
        plot_multi_qvalue_vs_state(
            {label: table},
            q_path,
            best_switch_points={label: best_sp_mab.get(label)},
            show_titles=False,
            show_legend=False,
            tick_fontsize=qvalue_tick_fontsize,
            state_min=state_min,
            state_max=state_max,
            state_step=state_step,
            x_tick_rotation=0,
        )

    for label, table in qvalue_tables_mc.items():
        q_path = os.path.join(qvalues_dir, f"{label.replace(' ', '_').lower()}_qvalues.{plot_format}")
        weights = [state for (state, _) in table.keys()]
        state_min = min(weights) if weights else None
        state_max = max(weights) if weights else None
        state_step = _infer_tick_step(weights)
        plot_multi_qvalue_pair_tables(
            {label: table},
            q_path,
            best_switch_points={label: best_sp_mc.get(label)},
            show_titles=False,
            show_legend=True,
            tick_fontsize=qvalue_tick_fontsize,
            legend_fontsize=max(24, qvalue_tick_fontsize),
            x_tick_rotation=0,
            force_last_xtick=False,
            state_min=state_min,
            state_max=state_max,
            state_step=state_step,
        )

    metrics_path = os.path.join(output_dir, "eval_compare_test_metrics.json")
    with open(metrics_path, "w", encoding="utf-8") as handle:
        json.dump(
            {
                "timestamp": timestamp,
                "artifacts": {
                    "compare_switching": compare_path,
                    "qvalues_dir": qvalues_dir,
                },
                "runs": {
                    "mab": mab_runs,
                    "mc": mc_runs,
                },
            },
            handle,
            indent=2,
        )

    logger.info("Saved output comparison plots to %s", output_dir)


if __name__ == "__main__":
    main()
