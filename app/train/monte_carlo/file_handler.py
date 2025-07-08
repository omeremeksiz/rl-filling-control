import pandas as pd
import webbrowser
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class FileHandler:
    def read_excel(self, excel_file_path):
        if not excel_file_path:
            logging.info("Excel file path not provided.")
            return None
        try:
            return pd.read_excel(excel_file_path, header=None)
        except Exception as e:
            logging.info(f"Error reading excel file: {e}")
            return None

    def write_to_text(self, file_path, header, data):
        try:
            with open(file_path, "w") as f:
                print(header, file=f)
                for item in data:
                    print(item, file=f)
        except Exception as e:
            logging.info(f"Error writing to text file: {e}")

    def append_to_text(self, file_path, data):
        try:
            with open(file_path, "a") as f:
                for item in data:
                    print(item, file=f)
        except Exception as e:
            logging.info(f"Error appending to text file: {e}")

    def read_txt(self, file_path, column_names=None):
        try:
            data = pd.read_csv(file_path, delimiter='\t', skiprows=1, header=None)
            if column_names:
                data.columns = column_names
            else:
                data.columns = ['Weight', 'Action', 'Count', 'Q Value']
            return data
        except Exception as e:
            logging.info(f"Error reading txt file: {e}")
            return None
        
    def write_to_excel(self, data, file_path):
        if data is not None:
            try:
                df = pd.DataFrame(data, columns=['Received Data'])
                df.to_excel(file_path, index=False)
            except Exception as e:
                logging.info(f"Error writing to excel file: {e}")
        else:
            logging.info("No data to write to Excel.")

    def write_tolerance_pairs(self, file_path, tolerance_pairs):
        sorted_pairs = sorted(tolerance_pairs.items(), key=lambda x: x[0].cutoff_weight)
        
        tolerance_output = []
        for pair, weight_margin in sorted_pairs:
            tolerance_output.append(f"{pair.cutoff_weight}\t{pair.error_type}\t{weight_margin}")

        header = "Cutoff Weight\tError Type\tWeight Margin"
        self.write_to_text(file_path, header, tolerance_output)

    def write_final_arr(self, file_path, final_arr):
        final_output = []

        for state in final_arr[5, :]:
            weight, action, reward = state
            final_output.append(f"{weight}\t{int(action)}\t{reward}")
        
        header = "Weight\tAction\tG Value"
        self.write_to_text(file_path, header, final_output)

    def write_q_values(self, file_path, q_values, count):
        q_value_output = []
        sorted_items = sorted(q_values.items(), key=lambda x: (x[0].weight, x[0].action))
        for state, reward in sorted_items:
            count_value = count[state]
            q_value_output.append(f"{state.weight}\t{int(state.action)}\t{count_value}\t{reward}")
        
        header = "Weight\tAction\tCount\tQ Value"
        self.write_to_text(file_path, header, q_value_output)

    def write_target_state_qvalues(self, file_path, q_values_for_target_state):
        q_value_output = []
        for i, (q_value, difference, reward) in enumerate(q_values_for_target_state, start=1):
            q_value_output.append(f"{i}\t{q_value}\t{difference}\t{reward}")

        header = "Iteration\tQ Value\tDifference\tG Value"
        self.write_to_text(file_path, header, q_value_output)

    def write_best_actions(self, file_path, best_action):
        best_actions_output = []
        for weight, (action, q_value, count_value) in best_action.items():
            best_actions_output.append(f"{weight}\t{action}\t{count_value}\t{q_value}")
        
        header = "Weight\tBest Action\tCount\tQ Value"
        self.write_to_text(file_path, header, best_actions_output)

    def format_and_write_table(self, input_file_path, output_html_path, column_names=None):
        df = self.read_txt(input_file_path, column_names)
        if df is not None:
            try:
                df.to_html(output_html_path, index=False, border=1, justify='center')
                self.open_in_browser(output_html_path)
            except Exception as e:
                logging.info(f"Error writing HTML file: {e}")
        else:
            logging.info("Failed to read and format the table. Please check the input file.")

    def merge_qvalue_files(self, mc_file_path, td_file_path, ql_file_path, output_html_path):
        try:
            mc_df = pd.read_csv(mc_file_path, delimiter='\t', skiprows=1, header=None, names=['Weight', 'Action', 'MC Count', 'MC Q Value'])
            td_df = pd.read_csv(td_file_path, delimiter='\t', skiprows=1, header=None, names=['Weight', 'Action', 'TD Count', 'TD Q Value'])
            ql_df = pd.read_csv(ql_file_path, delimiter='\t', skiprows=1, header=None, names=['Weight', 'Action', 'QL Count', 'QL Q Value'])

            merged_df = pd.merge(mc_df, td_df, on=['Weight', 'Action'])
            merged_df = pd.merge(merged_df, ql_df, on=['Weight', 'Action'])
            merged_df = merged_df[['Weight', 'Action', 'MC Count', 'MC Q Value', 'TD Count', 'TD Q Value', 'QL Count', 'QL Q Value']]
            
            merged_df.to_html(output_html_path, index=False, border=1, justify='center')
            self.open_in_browser(output_html_path)
        except Exception as e:
            logging.info(f"Error merging Q-value files: {e}")

    def merged_best_actions(self, mc_file_path, td_file_path, ql_file_path, output_html_path):
        try:
            mc_df = self.read_txt(mc_file_path, column_names=['Weight', 'MC Best Action', 'MC Count', 'MC Q Value'])
            td_df = self.read_txt(td_file_path, column_names=['Weight', 'TD Best Action', 'TD Count', 'TD Q Value'])
            ql_df = self.read_txt(ql_file_path, column_names=['Weight', 'QL Best Action', 'QL Count', 'QL Q Value'])

            if mc_df is None or td_df is None or ql_df is None:
                logging.info("Failed to read input files. Please check the file paths.")
                return

            merged_df = pd.merge(mc_df, td_df, on='Weight')
            merged_df = pd.merge(merged_df, ql_df, on='Weight')
            merged_df = merged_df[['Weight', 'MC Best Action', 'MC Count', 'MC Q Value', 'TD Best Action', 'TD Count', 'TD Q Value', 'QL Best Action', 'QL Count', 'QL Q Value']]
            
            merged_df.to_html(output_html_path, index=False, border=1, justify='center')
            self.open_in_browser(output_html_path)
        except Exception as e:
            logging.info(f"Error merging best actions files: {e}")

    def write_qvalue_updates_to_excel(self, all_episodes_info, file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "qvalue_updates"

        # Styles
        font_bold = Font(bold=True, size=12)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        thick_side = Side(style='thick', color='000000')

        # Fill colors
        light_gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        dark_gray_fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Normal
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Underflow
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Overflow

        start_col = 1

        for episode_info in all_episodes_info:
            episode_col = start_col

            # Merge for Episode title
            ws.merge_cells(start_row=1, start_column=episode_col, end_row=1, end_column=episode_col + 3)
            title_cell = ws.cell(row=1, column=episode_col)
            title_cell.value = f"Episode {episode_info['episode_num'] + 1}"
            title_cell.font = font_bold
            title_cell.alignment = center_alignment
            title_cell.fill = light_gray_fill

            termination_type = episode_info.get('termination_type', 'Normal')  

            headers = [
                f"Experienced: {episode_info['experienced_switching_point']}",
                f"Model-Selected: {episode_info['model_selected_switching_point']}",
                f"Explored: {episode_info['explored_switching_point'] if episode_info['explored_switching_point'] is not None else 'None'}",
                f"Result: {termination_type}"  
            ]
            
            for idx, header in enumerate(headers):
                cell = ws.cell(row=2, column=episode_col + idx)
                cell.value = header
                cell.font = font_bold
                cell.alignment = center_alignment

                if idx == 3:
                    termination_fill = {
                        "Normal": green_fill,
                        "Underflow": yellow_fill,
                        "Overflow": red_fill
                    }.get(termination_type, green_fill)
                    cell.fill = termination_fill

            # Column titles
            columns = ["Weight", "Action", "Count", "Q Value"]
            fills = [light_gray_fill, dark_gray_fill, light_gray_fill, dark_gray_fill]
            for idx, (col_name, fill) in enumerate(zip(columns, fills)):
                cell = ws.cell(row=3, column=episode_col + idx)
                cell.value = col_name
                cell.font = font_bold
                cell.fill = fill
                cell.alignment = center_alignment

            # Write q-value data
            row_idx = 4
            experienced_weight = episode_info['experienced_switching_point']

            for state, q_val in episode_info['qValue'].items():
                weight = state.weight
                action = state.action
                count_val = episode_info['count'][state]

                # Check if this is the experienced weight
                is_experienced_row = (weight == experienced_weight)

                ws.cell(row=row_idx, column=episode_col, value=weight).font = font_bold
                ws.cell(row=row_idx, column=episode_col + 1, value=action).font = font_bold
                ws.cell(row=row_idx, column=episode_col + 2, value=count_val).font = font_bold

                q_value_cell = ws.cell(row=row_idx, column=episode_col + 3)
                q_value_cell.value = float(f"{q_val:.4f}")
                q_value_cell.font = font_bold

                # If experienced weight, highlight the full row
                if is_experienced_row:
                    for offset in range(4):
                        ws.cell(row=row_idx, column=episode_col + offset).fill = highlight_fill

                row_idx += 1

            # Adjust column widths
            for idx in range(4):
                col_letter = get_column_letter(episode_col + idx)
                ws.column_dimensions[col_letter].width = 18

            # Apply borders
            top_row = 1
            bottom_row = row_idx - 1

            for row in range(top_row, bottom_row + 1):
                for col in range(episode_col, episode_col + 4):
                    cell = ws.cell(row=row, column=col)
                    # Set thin borders
                    cell.border = thin_border

                    # Apply thick outer border
                    if row == top_row:
                        cell.border = Border(top=thick_side, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
                    if row == bottom_row:
                        cell.border = Border(bottom=thick_side, left=cell.border.left, right=cell.border.right, top=cell.border.top)
                    if col == episode_col:
                        cell.border = Border(left=thick_side, top=cell.border.top, bottom=cell.border.bottom, right=cell.border.right)
                    if col == episode_col + 3:
                        cell.border = Border(right=thick_side, top=cell.border.top, bottom=cell.border.bottom, left=cell.border.left)

            # Move to next block
            start_col += 4

        wb.save(file_path)

    def open_in_browser(self, file_path):
        try:
            if os.path.exists(file_path):
                webbrowser.open(f'file://{os.path.realpath(file_path)}', new=2)
            else:
                logging.info("HTML file not found.")
        except Exception as e:
            logging.info(f"Error opening HTML file in browser: {e}")
