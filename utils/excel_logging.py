# utils/excel_logging.py
from __future__ import annotations

from typing import Mapping, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_FONT_BOLD = Font(bold=True, size=12)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
_THICK_SIDE = Side(style="thick", color="000000")

_FILL_LIGHT_GRAY = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
_FILL_DARK_GRAY = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
_FILL_HIGHLIGHT = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_FILL_TERMINATION = {
    "normal": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "safe": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "underflow": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "underweight": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "overflow": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "overweight": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}


def write_mab_qtable_to_excel(
    all_episodes_info: Sequence[Mapping[str, object]],
    file_path: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "mab_qvalue_updates"

    start_col = 1
    column_widths = (20, 20, 26)

    for episode_info in all_episodes_info:
        episode_col = start_col
        episode_num = int(episode_info["episode_num"]) + 1
        experienced = episode_info["experienced_switching_point"]
        model_selected = episode_info.get("model_selected_switching_point")
        explored = episode_info.get("explored_switching_point")
        termination_type = str(episode_info.get("termination_type", "normal")).lower()
        q_table: Mapping[int, float] = episode_info.get("q_table", {})  # type: ignore[assignment]
        counts: Mapping[int, int] = episode_info.get("counts", {})  # type: ignore[assignment]

        # Episode title row
        ws.merge_cells(
            start_row=1,
            start_column=episode_col,
            end_row=1,
            end_column=episode_col + 2,
        )
        title_cell = ws.cell(row=1, column=episode_col)
        title_cell.value = f"Episode {episode_num}"
        title_cell.font = _FONT_BOLD
        title_cell.alignment = _ALIGN_CENTER
        title_cell.fill = _FILL_LIGHT_GRAY

        meta_cells = (
            (f"Experienced: {experienced}", episode_col, None),
            (f"Model-Selected: {model_selected}", episode_col + 1, None),
            (
                f"Explored: {explored if explored is not None else 'None'}\n"
                f"Result: {termination_type.capitalize()}",
                episode_col + 2,
                _FILL_TERMINATION.get(termination_type, _FILL_TERMINATION["normal"]),
            ),
        )

        for text, column, fill in meta_cells:
            cell = ws.cell(row=2, column=column)
            cell.value = text
            cell.font = _FONT_BOLD
            cell.alignment = _ALIGN_CENTER_WRAP if "\n" in text else _ALIGN_CENTER
            if fill:
                cell.fill = fill

        column_titles = ("Switch Point", "Updates", "Q Value")
        fills = (_FILL_LIGHT_GRAY, _FILL_DARK_GRAY, _FILL_LIGHT_GRAY)
        for idx, (title, fill) in enumerate(zip(column_titles, fills)):
            cell = ws.cell(row=3, column=episode_col + idx)
            cell.value = title
            cell.font = _FONT_BOLD
            cell.fill = fill
            cell.alignment = _ALIGN_CENTER

        # rows of switch points sorted for readability
        sorted_switch_points = sorted(q_table.keys())
        for row_idx, sp in enumerate(sorted_switch_points, start=4):
            ws.cell(row=row_idx, column=episode_col, value=sp).font = _FONT_BOLD
            ws.cell(
                row=row_idx,
                column=episode_col + 1,
                value=counts.get(sp, 0),
            ).font = _FONT_BOLD
            q_value_cell = ws.cell(
                row=row_idx,
                column=episode_col + 2,
                value=_format_float(q_table.get(sp, 0.0)),
            )
            q_value_cell.font = _FONT_BOLD

            if sp == experienced:
                for offset in range(3):
                    ws.cell(row=row_idx, column=episode_col + offset).fill = _FILL_HIGHLIGHT

        bottom_row = 3 + len(sorted_switch_points)
        _apply_borders(ws, episode_col, episode_col + 2, 1, bottom_row)

        for idx, width in enumerate(column_widths):
            col_letter = get_column_letter(episode_col + idx)
            ws.column_dimensions[col_letter].width = width

        start_col += 3

    wb.save(file_path)


def write_qtable_to_excel(
    all_episodes_info: Sequence[Mapping[str, object]],
    file_path: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "qvalue_updates"

    start_col = 1
    column_widths = (16, 16, 16, 20)

    for episode_info in all_episodes_info:
        episode_col = start_col
        episode_num = int(episode_info["episode_num"]) + 1
        experienced = episode_info["experienced_switching_point"]
        model_selected = episode_info.get("model_selected_switching_point")
        explored = episode_info.get("explored_switching_point")
        termination_type = str(episode_info.get("termination_type", "normal")).lower()
        q_table: Mapping[Tuple[int, int], float] = episode_info.get("q_table", {})  # type: ignore[assignment]
        counts: Mapping[Tuple[int, int], int] = episode_info.get("counts", {})  # type: ignore[assignment]

        ws.merge_cells(
            start_row=1,
            start_column=episode_col,
            end_row=1,
            end_column=episode_col + 3,
        )
        title_cell = ws.cell(row=1, column=episode_col)
        title_cell.value = f"Episode {episode_num}"
        title_cell.font = _FONT_BOLD
        title_cell.alignment = _ALIGN_CENTER
        title_cell.fill = _FILL_LIGHT_GRAY

        meta_cells = (
            (f"Experienced: {experienced}", episode_col, None),
            (f"Model-Selected: {model_selected}", episode_col + 1, None),
            (f"Explored: {explored if explored is not None else 'None'}", episode_col + 2, None),
            (
                f"Result: {termination_type.capitalize()}",
                episode_col + 3,
                _FILL_TERMINATION.get(termination_type, _FILL_TERMINATION["normal"]),
            ),
        )

        for text, column, fill in meta_cells:
            cell = ws.cell(row=2, column=column)
            cell.value = text
            cell.font = _FONT_BOLD
            cell.alignment = _ALIGN_CENTER
            if fill:
                cell.fill = fill

        column_titles = ("Weight", "Action", "Updates", "Q Value")
        fills = (_FILL_LIGHT_GRAY, _FILL_DARK_GRAY, _FILL_LIGHT_GRAY, _FILL_DARK_GRAY)
        for idx, (title, fill) in enumerate(zip(column_titles, fills)):
            cell = ws.cell(row=3, column=episode_col + idx)
            cell.value = title
            cell.font = _FONT_BOLD
            cell.fill = fill
            cell.alignment = _ALIGN_CENTER

        sorted_pairs = sorted(q_table.keys(), key=lambda pair: (pair[0], -pair[1]))
        for row_idx, (weight, action) in enumerate(sorted_pairs, start=4):
            ws.cell(row=row_idx, column=episode_col, value=weight).font = _FONT_BOLD
            ws.cell(row=row_idx, column=episode_col + 1, value=action).font = _FONT_BOLD
            ws.cell(
                row=row_idx,
                column=episode_col + 2,
                value=counts.get((weight, action), 0),
            ).font = _FONT_BOLD
            q_value_cell = ws.cell(
                row=row_idx,
                column=episode_col + 3,
                value=_format_float(q_table.get((weight, action), 0.0)),
            )
            q_value_cell.font = _FONT_BOLD

            if weight == experienced:
                for offset in range(4):
                    ws.cell(row=row_idx, column=episode_col + offset).fill = _FILL_HIGHLIGHT

        bottom_row = 3 + len(sorted_pairs)
        _apply_borders(ws, episode_col, episode_col + 3, 1, bottom_row)

        for idx, width in enumerate(column_widths):
            col_letter = get_column_letter(episode_col + idx)
            ws.column_dimensions[col_letter].width = width

        start_col += 4

    wb.save(file_path)


def _apply_borders(ws, start_col: int, end_col: int, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = _THIN_BORDER
            if row == start_row:
                cell.border = Border(
                    top=_THICK_SIDE,
                    left=cell.border.left,
                    right=cell.border.right,
                    bottom=cell.border.bottom,
                )
            if row == end_row:
                cell.border = Border(
                    bottom=_THICK_SIDE,
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                )
            if col == start_col:
                cell.border = Border(
                    left=_THICK_SIDE,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                    right=cell.border.right,
                )
            if col == end_col:
                cell.border = Border(
                    right=_THICK_SIDE,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                    left=cell.border.left,
                )


def _format_float(value: float) -> float:
    return float(f"{value:.4f}")
